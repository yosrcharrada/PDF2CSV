"""Writing results out, with the validation report attached — always.

There is no code path in this project that writes a CSV without also writing
its report. That is the whole point: a CSV found on a shared drive three weeks
later still carries proof of whether it reconciled, and nobody has to remember
whether this one was checked.

Failed checks do not block the write. They are printed loudly, logged, and
recorded in the sidecar. An analyst who cannot get their data out will export
it some other way and lose the report entirely, which is the outcome this
design is trying to avoid.
"""

from __future__ import annotations

import json
from dataclasses import dataclass, field
from pathlib import Path

import pandas as pd

from pdf2csv.logging_setup import get_logger
from pdf2csv.models import ExtractionResult, Severity
from pdf2csv.wording import count

log = get_logger(__name__)

# Excel on Windows reads a bare UTF-8 CSV as the system codepage, turning
# "Débit" into "DÃ©bit". The BOM makes it read UTF-8 correctly on a double
# click, which is exactly how the analyst is going to open this file.
CSV_ENCODING = "utf-8-sig"

_ERROR_FILL = "FFF4CCCC"
_WARNING_FILL = "FFFFF2CC"
_HEADER_FILL = "FF1F3864"


@dataclass
class ExportPaths:
    """Where everything ended up. Returned so callers can show or open them."""

    csv: Path
    report_json: Path
    xlsx: Path | None = None
    extras: list[Path] = field(default_factory=list)

    def all_paths(self) -> list[Path]:
        return [p for p in (self.csv, self.report_json, self.xlsx) if p] + self.extras


def export_result(
    result: ExtractionResult,
    path: str | Path,
    *,
    write_xlsx: bool = True,
    write_extras: bool = True,
) -> ExportPaths:
    """Write the CSV, its validation sidecar, and optionally an Excel workbook.

    Args:
        result: What :func:`pdf2csv.run` returned.
        path: Destination for the CSV. Sidecar and workbook are named from it.
        write_xlsx: Also produce a formatted ``.xlsx`` with flagged cells
            highlighted and the validation report on its own sheet.
        write_extras: Write any secondary tables as ``<name>.table2.csv`` etc.

    Returns:
        The paths written.
    """
    csv_path = Path(path)
    csv_path.parent.mkdir(parents=True, exist_ok=True)

    frame: pd.DataFrame = result.dataframe
    frame.to_csv(csv_path, index=False, encoding=CSV_ENCODING)

    report_path = csv_path.with_suffix(".validation.json")
    report_path.write_text(
        json.dumps(result.to_dict(), indent=2, ensure_ascii=False), encoding="utf-8"
    )

    paths = ExportPaths(csv=csv_path, report_json=report_path)

    if write_extras and result.extra_frames:
        for index, extra in enumerate(result.extra_frames, start=2):
            extra_path = csv_path.with_suffix(f".table{index}.csv")
            extra.to_csv(extra_path, index=False, encoding=CSV_ENCODING)
            paths.extras.append(extra_path)

    if write_xlsx:
        try:
            paths.xlsx = write_workbook(result, csv_path.with_suffix(".xlsx"))
        except Exception as exc:
            # The CSV is the deliverable; a workbook failure must not sink it.
            log.warning("could not write the Excel workbook: %s", exc)

    _announce(result, paths)
    return paths


def _announce(result: ExtractionResult, paths: ExportPaths) -> None:
    """Say plainly what happened. Stdout matters: the notebook path reads it."""
    report = result.report
    failed = [c for c in report.failed_checks if c.severity is Severity.ERROR]
    warned = [c for c in report.failed_checks if c.severity is Severity.WARNING]

    if failed:
        print(
            f"\n  {count(len(failed), 'check')} did not pass "
            "— review this before using the file:"
        )
        for check in failed:
            print(f"    - {check.title}: {check.detail}")
    if warned:
        print(f"\n  {count(len(warned), 'warning')}:")
        for check in warned:
            print(f"    - {check.title}: {check.detail}")
    if not failed and not warned:
        print(f"\n  All {len(report.checks)} checks passed.")

    print(f"\n  CSV     {paths.csv}")
    print(f"  Report  {paths.report_json}")
    if paths.xlsx:
        print(f"  Excel   {paths.xlsx}")
    for extra in paths.extras:
        print(f"  Extra   {extra}")
    print()

    log.info("exported %s — %s", paths.csv.name, report.summary())


# --------------------------------------------------------------------------- #
# Excel workbook
# --------------------------------------------------------------------------- #


def write_workbook(result: ExtractionResult, path: Path) -> Path:
    """Write a formatted workbook: data, highlighted flags, and the report.

    Worth the extra code because it closes the loop for the analyst. The report
    says "row 42 is suspicious"; in the workbook, row 42 is already yellow. No
    cross-referencing a JSON file against a spreadsheet by hand.
    """
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    frame: pd.DataFrame = result.dataframe

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        frame.to_excel(writer, sheet_name="Data", index=False)
        _write_validation_sheet(writer, result)
        _write_document_sheet(writer, result)

        for index, extra in enumerate(result.extra_frames, start=2):
            extra.to_excel(writer, sheet_name=f"Table {index}", index=False)

        workbook = writer.book
        sheet = workbook["Data"]

        # Header styling and a frozen top row: a 400-row ledger is unusable
        # without it, and the analyst should not have to do it themselves.
        header_font = Font(bold=True, color="FFFFFFFF")
        header_fill = PatternFill("solid", fgColor=_HEADER_FILL)
        for cell in sheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(vertical="center")
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions

        column_index = {str(name): i + 1 for i, name in enumerate(frame.columns)}

        # Number format for amount columns, so 1234.5 reads as 1,234.50.
        for name, position in column_index.items():
            if pd.api.types.is_numeric_dtype(frame[name]):
                letter = get_column_letter(position)
                for cell in sheet[letter][1:]:
                    cell.number_format = "#,##0.00"

        _apply_flag_highlighting(sheet, result, column_index)
        _autosize(sheet, frame)

    log.info("wrote workbook %s", path.name)
    return path


def _apply_flag_highlighting(sheet, result: ExtractionResult, column_index: dict[str, int]) -> None:
    from openpyxl.comments import Comment
    from openpyxl.styles import PatternFill

    error_fill = PatternFill("solid", fgColor=_ERROR_FILL)
    warning_fill = PatternFill("solid", fgColor=_WARNING_FILL)

    for flag in result.report.flags:
        position = column_index.get(flag.column)
        if position is None:
            continue
        # +2: one for the header row, one because openpyxl is 1-based.
        cell = sheet.cell(row=flag.row + 2, column=position)
        cell.fill = error_fill if flag.severity is Severity.ERROR else warning_fill
        note = flag.reason
        if flag.value:
            note += f"\n\nPDF showed: {flag.value}"
        cell.comment = Comment(note, "PDF2CSV")


def _write_validation_sheet(writer, result: ExtractionResult) -> None:
    checks = pd.DataFrame(
        [
            {
                "Check": c.title,
                "Result": "PASS" if c.passed else c.severity.value.upper(),
                "Detail": c.detail,
                "What to do": c.hint,
            }
            for c in result.report.checks
        ]
    )
    if checks.empty:
        checks = pd.DataFrame(
            [{"Check": "No checks ran", "Result": "", "Detail": "", "What to do": ""}]
        )
    checks.to_excel(writer, sheet_name="Validation", index=False)


def _write_document_sheet(writer, result: ExtractionResult) -> None:
    meta = result.meta
    rows = [
        ("Source file", meta.source_name),
        ("SHA-256", meta.sha256),
        ("Pages", meta.n_pages),
        ("Digital pages", meta.n_digital),
        ("Scanned pages", meta.n_scanned),
        ("Document profile", meta.profile),
        ("Rows extracted", result.n_rows),
        ("Processing time", f"{meta.duration_seconds:.1f} s"),
        ("Overall result", "PASSED" if result.report.passed else "REVIEW REQUIRED"),
        ("Generated", result.report.generated_at),
    ]
    rows.extend(("Warning", w) for w in meta.warnings)
    pd.DataFrame(rows, columns=["Field", "Value"]).to_excel(
        writer, sheet_name="Document", index=False
    )


def _autosize(sheet, frame: pd.DataFrame) -> None:
    """Approximate column widths from content length, capped so one long
    description does not push every other column off the screen."""
    from openpyxl.utils import get_column_letter

    for position, name in enumerate(frame.columns, start=1):
        longest = max(
            [len(str(name))] + [len(str(v)) for v in frame[name].head(200).tolist()],
            default=10,
        )
        sheet.column_dimensions[get_column_letter(position)].width = min(
            max(longest + 2, 10), 50
        )


# --------------------------------------------------------------------------- #
# Notebook-facing convenience
# --------------------------------------------------------------------------- #


def export(result: ExtractionResult, path: str | Path = "output.csv") -> Path:
    """One-liner for the notebook template.

    Exists so the notebook's third cell is a single call with no arguments to
    get wrong, and so validation cannot be skipped by not running a cell —
    there is no way to write a CSV from the notebook that does not pass through
    here.
    """
    return export_result(result, path).csv
